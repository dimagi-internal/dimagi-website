# Regenerates "Dimagi Blog Tags.xlsx" from the CURRENT live blog + the consolidation pass.
# Per blog: Focus, Type (unchanged) + a single specific COUNTRY and a single Solutions-aligned
# TOPIC tag (sector > use case > org type) replacing the old continent + SEO1 + SEO2.
# Mapping logic lives in consolidation_maps.py (shared with apply_tag_pass.py).
#   python3 Context/blog-import/build_tag_xlsx.py
import re, os, sys, json, datetime
from collections import Counter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from consolidation_maps import resolve_country, resolve_topic, is_skip

ROOT = "/Users/gillianjavetski/Documents/Gillian Coding/Pre-Login Websites/Dimagi Pre-Login"
OUT  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Dimagi Blog Tags.xlsx")
IDX  = os.path.join(ROOT, "blog", "index.html")
seo  = json.load(open(os.path.join(ROOT, "Context/blog-import/seo_tags.json")))

def txt(s): return re.sub(r'<[^>]+>', '', s).strip()

def pub_date(slug):
    fp = os.path.join(ROOT, "blog", slug, "index.html")
    if not os.path.exists(fp): return None
    h = open(fp, encoding='utf-8').read()
    m = re.search(r'"datePublished":\s*"(\d{4})-(\d{2})-(\d{2})"', h) \
        or re.search(r'article:published_time"\s+content="(\d{4})-(\d{2})-(\d{2})"', h)
    return datetime.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

idx = open(IDX, encoding='utf-8').read()
rows = []
for block in re.findall(r'<article class="blog-card".*?</article>', idx, flags=re.S):
    a = dict(re.findall(r'data-(\w+)="([^"]*)"', re.search(r'<article[^>]*>', block).group(0)))
    slug = re.search(r'href="([^"/]+)/index\.html"', block).group(1)
    title = txt(re.search(r'<h2 class="blog-card-title">(.*?)</h2>', block, re.S).group(1))
    dm = re.search(r'<p class="blog-card-desc">(.*?)</p>', block, re.S)
    desc = txt(dm.group(1)) if dm else ''
    focus = a.get('product') if a.get('product') not in (None,'None','') else \
            (a.get('topic') if a.get('topic') not in (None,'None','') else 'Dimagi')
    ctype = a.get('type', '')
    geo = a.get('country', 'None'); geo = '' if geo in ('None','') else geo
    shown = txt(re.search(r'<span class="blog-card-date">(.*?)</span>', block, re.S).group(1))
    s = seo.get(slug, {}); s1, s2 = s.get('seo1',''), s.get('seo2','')
    rows.append({'pub': pub_date(slug), 'shown': shown, 'title': title,
                 'url': f'https://dimagi.com/{slug}/', 'focus': focus, 'type': ctype,
                 'country': resolve_country(slug, geo, s1, s2, title),
                 'topic': resolve_topic(slug, s1, s2),
                 'old_geo': geo, 's1': s1, 's2': s2, 'slug': slug, 'desc': desc})

rows.sort(key=lambda r: r['pub'] or datetime.datetime(1900,1,1), reverse=True)

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Blogs"
HEAD = ['Published Date','Date Shown','Blog Name','Blog URL','Focus','Type',
        'Country','Solutions Tag','— old Geography','— old SEO 1','— old SEO 2',
        'Slug','Description']
ws.append(HEAD)
for r in rows:
    ws.append([r['pub'], r['shown'], r['title'], r['url'], r['focus'], r['type'],
               r['country'], r['topic'], r['old_geo'], r['s1'], r['s2'], r['slug'], r['desc']])

blue = PatternFill("solid", fgColor="3843D0"); grey = PatternFill("solid", fgColor="9AA0B4")
for i, c in enumerate(ws[1], 1):
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = grey if HEAD[i-1].startswith('—') else blue
    c.alignment = Alignment(vertical="center", wrap_text=True)
for r in range(2, ws.max_row+1):
    ws.cell(r,1).number_format = 'yyyy-mm-dd'
    ws.cell(r,4).hyperlink = ws.cell(r,4).value; ws.cell(r,4).font = Font(color="0563C1", underline="single")
    for col in (9,10,11):
        ws.cell(r,col).font = Font(color="8A8F9E")
widths = [13,10,46,40,15,12,16,30,13,16,16,38,60]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{ws.max_row}"

vs = wb.create_sheet("Tag Vocabulary")
vs.append(["New tagging: Focus + Type are the blog filters. Country + Solutions Tag are page-only "
           "(shown in 'Filed under'). Solutions Tag aligns with dimagi.com Solutions sectors/use-cases/org-types."])
vs["A1"].font = Font(bold=True, italic=True)
for dim,key in [("Focus",'focus'),("Type",'type'),("Solutions Tag",'topic'),("Country",'country')]:
    vs.append([]); vs.append([dim,"Count"])
    vs.cell(vs.max_row,1).font = Font(bold=True); vs.cell(vs.max_row,2).font = Font(bold=True)
    for val,n in sorted(Counter(r[key] or '(none)' for r in rows).items(), key=lambda x:-x[1]):
        vs.append([val,n])
vs.column_dimensions['A'].width = 32; vs.column_dimensions['B'].width = 8

wb.save(OUT)
print(f"wrote {OUT}  ({len(rows)} blogs)")
print("Solutions Tag:", dict(Counter(r['topic'] or '(none)' for r in rows)))
print("Country      :", dict(Counter(r['country'] or '(none)' for r in rows)))
